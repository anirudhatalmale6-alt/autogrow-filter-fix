#!/usr/bin/env python3
"""
Analyze virtual screening results: Top 500 by Vina Score and Ligand Efficiency.

Reads docking results, merges with NP Atlas / COCONUT metadata,
and generates an Excel report with two sheets:
  1. Top 500 by Vina Score (most negative = best)
  2. Top 500 by Ligand Efficiency (LE = -score / heavy_atom_count)

Usage:
  python3 analyze_screening_results.py <results_file> [--metadata <metadata.tsv>] [--top N]

Results file can be:
  - TSV/CSV with columns: name/compound_id, smiles, score/vina_score
  - AutoGrow4 output ranked_compound_diversity.tsv or similar
  - Vina output log parsed into tabular format

If no metadata file given, searches ~/final_project/np_databases/ automatically.
"""
import argparse
import csv
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from rdkit import Chem
    from rdkit.Chem import Descriptors
    HAS_RDKIT = True
except ImportError:
    HAS_RDKIT = False


def load_metadata(meta_paths):
    """Load metadata from one or more TSV files into a dict keyed by compound_id."""
    meta = {}
    for path in meta_paths:
        if not os.path.exists(path):
            continue
        print(f"  Loading metadata: {path}")
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.DictReader(f, delimiter="\t")
            for row in reader:
                cid = row.get("compound_id", "").strip()
                if cid:
                    meta[cid] = row
    print(f"  Total metadata entries: {len(meta)}")
    return meta


def count_heavy_atoms_from_smiles(smi):
    """Count heavy atoms from SMILES string."""
    if HAS_RDKIT:
        mol = Chem.MolFromSmiles(smi)
        if mol:
            return mol.GetNumHeavyAtoms()
    count = 0
    i = 0
    while i < len(smi):
        c = smi[i]
        if c in "([])+-=#:/\\.@0123456789 \t":
            i += 1
            continue
        if c.isupper():
            if c == 'H':
                if i + 1 < len(smi) and smi[i + 1].islower():
                    count += 1
                    i += 2
                    continue
                i += 1
                continue
            count += 1
            if i + 1 < len(smi) and smi[i + 1].islower():
                i += 2
            else:
                i += 1
        else:
            i += 1
    return max(count, 1)


def load_results(results_path):
    """Load screening results from various formats."""
    results = []
    with open(results_path, "r", encoding="utf-8", errors="replace") as f:
        first_line = f.readline().strip()
        f.seek(0)

        sep = "\t" if "\t" in first_line else ","
        reader = csv.DictReader(f, delimiter=sep)
        fields = reader.fieldnames
        if fields is None:
            print("ERROR: Could not read header from results file")
            return []

        fields_lower = [fld.lower().strip() for fld in fields]

        name_col = None
        for candidate in ["name", "compound_id", "ligand", "id", "npaid", "identifier", "mol_name"]:
            if candidate in fields_lower:
                name_col = fields[fields_lower.index(candidate)]
                break
        if name_col is None:
            name_col = fields[0]

        smiles_col = None
        for candidate in ["smiles", "compound_smiles", "canonical_smiles", "smi"]:
            if candidate in fields_lower:
                smiles_col = fields[fields_lower.index(candidate)]
                break

        score_col = None
        for candidate in ["score", "vina_score", "docking_score", "affinity", "fitness"]:
            if candidate in fields_lower:
                score_col = fields[fields_lower.index(candidate)]
                break
        if score_col is None:
            for i, fld in enumerate(fields_lower):
                if "score" in fld or "affinity" in fld or "fitness" in fld:
                    score_col = fields[i]
                    break

        status_col = None
        for candidate in ["status", "result"]:
            if candidate in fields_lower:
                status_col = fields[fields_lower.index(candidate)]
                break

        print(f"  Detected columns - Name: {name_col}, SMILES: {smiles_col}, Score: {score_col}")

        for row in reader:
            if status_col and row.get(status_col, "").strip().lower() not in ["ok", "success", ""]:
                continue

            try:
                score = float(row.get(score_col, "nan"))
            except (ValueError, TypeError):
                continue

            if score != score and score == float("nan"):
                continue

            name = row.get(name_col, "unknown").strip()
            smiles = row.get(smiles_col, "") if smiles_col else ""

            results.append({
                "name": name,
                "smiles": smiles.strip(),
                "score": score,
            })

    print(f"  Loaded {len(results)} valid results")
    return results


def compute_ligand_efficiency(results):
    """Add heavy atom count and ligand efficiency to results."""
    for r in results:
        smi = r.get("smiles", "")
        if smi:
            ha = count_heavy_atoms_from_smiles(smi)
        else:
            ha = 0
        r["heavy_atoms"] = ha
        if ha > 0 and r["score"] < 0:
            r["ligand_efficiency"] = -r["score"] / ha
        else:
            r["ligand_efficiency"] = 0.0


def write_excel(results_by_vina, results_by_le, metadata, output_path, top_n):
    """Write Excel workbook with two sheets."""
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    columns = [
        ("Rank", 6),
        ("Compound ID", 16),
        ("Compound Name", 30),
        ("SMILES", 50),
        ("Vina Score", 12),
        ("Heavy Atoms", 12),
        ("Ligand Efficiency", 16),
        ("MW", 10),
        ("Source Organism", 25),
        ("Chemical Class", 20),
        ("Database", 12),
    ]

    def write_sheet(ws, title, ranked_results):
        ws.title = title
        for col_idx, (col_name, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A" + chr(64 + col_idx - 26)].width = width

        for rank, r in enumerate(ranked_results[:top_n], 1):
            cid = r["name"]
            meta = metadata.get(cid, {})

            name = meta.get("name", "")
            mw = meta.get("mw", meta.get("molecular_weight", ""))
            organism = meta.get("organisms", "")
            if not organism:
                genus = meta.get("genus", "")
                species = meta.get("species", "")
                organism = f"{genus} {species}".strip()
            chem_class = meta.get("chemical_class", meta.get("np_class", ""))
            db = "NP Atlas" if cid.startswith("NPA") else "COCONUT" if cid.startswith("CNP") else "Unknown"

            row_data = [
                rank,
                cid,
                name,
                r["smiles"],
                round(r["score"], 2),
                r["heavy_atoms"],
                round(r["ligand_efficiency"], 4),
                mw,
                organism,
                chem_class,
                db,
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=rank + 1, column=col_idx, value=val)
                cell.border = thin_border
                if col_idx in (1, 5, 6, 7, 8):
                    cell.alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"

    ws_vina = wb.active
    write_sheet(ws_vina, f"Top {top_n} by Vina Score", ranked_results=results_by_vina)

    ws_le = wb.create_sheet()
    write_sheet(ws_le, f"Top {top_n} by Ligand Efficiency", ranked_results=results_by_le)

    wb.save(output_path)
    print(f"  Excel report saved: {output_path}")


def write_tsv_fallback(results_by_vina, results_by_le, metadata, output_dir, top_n):
    """Fallback if openpyxl not available: write TSV files."""
    for label, ranked in [("vina_score", results_by_vina), ("ligand_efficiency", results_by_le)]:
        path = os.path.join(output_dir, f"top_{top_n}_{label}.tsv")
        with open(path, "w") as f:
            f.write("rank\tcompound_id\tname\tsmiles\tvina_score\theavy_atoms\tligand_efficiency\tmw\torganism\tchemical_class\tdatabase\n")
            for rank, r in enumerate(ranked[:top_n], 1):
                cid = r["name"]
                meta = metadata.get(cid, {})
                name = meta.get("name", "")
                mw = meta.get("mw", meta.get("molecular_weight", ""))
                organism = meta.get("organisms", "")
                if not organism:
                    organism = f"{meta.get('genus', '')} {meta.get('species', '')}".strip()
                chem_class = meta.get("chemical_class", meta.get("np_class", ""))
                db = "NP Atlas" if cid.startswith("NPA") else "COCONUT" if cid.startswith("CNP") else "Unknown"
                f.write(f"{rank}\t{cid}\t{name}\t{r['smiles']}\t{r['score']:.2f}\t{r['heavy_atoms']}\t{r['ligand_efficiency']:.4f}\t{mw}\t{organism}\t{chem_class}\t{db}\n")
        print(f"  TSV report saved: {path}")


def main():
    parser = argparse.ArgumentParser(description="Analyze virtual screening results")
    parser.add_argument("results_file", help="Path to screening results (TSV or CSV)")
    parser.add_argument("--metadata", nargs="*", help="Path(s) to metadata TSV file(s)")
    parser.add_argument("--top", type=int, default=500, help="Number of top compounds (default: 500)")
    parser.add_argument("--output", help="Output Excel file path")
    args = parser.parse_args()

    print("=" * 60)
    print("Virtual Screening Results Analysis")
    print("=" * 60)

    if args.metadata:
        meta_paths = args.metadata
    else:
        np_dir = os.path.expanduser("~/final_project/np_databases")
        meta_paths = [
            os.path.join(np_dir, "npatlas_metadata.tsv"),
            os.path.join(np_dir, "coconut_metadata.tsv"),
        ]

    print("\n1. Loading metadata...")
    metadata = load_metadata(meta_paths)

    print("\n2. Loading screening results...")
    results = load_results(args.results_file)
    if not results:
        print("ERROR: No valid results found")
        sys.exit(1)

    print("\n3. Computing ligand efficiency...")
    compute_ligand_efficiency(results)

    results_by_vina = sorted(results, key=lambda x: x["score"])
    results_by_le = sorted(results, key=lambda x: -x["ligand_efficiency"])

    top_n = args.top
    print(f"\n4. Best Vina score: {results_by_vina[0]['score']:.2f} ({results_by_vina[0]['name']})")
    print(f"   Best LE: {results_by_le[0]['ligand_efficiency']:.4f} ({results_by_le[0]['name']})")

    output_dir = os.path.dirname(os.path.abspath(args.results_file))
    if args.output:
        output_path = args.output
    else:
        output_path = os.path.join(output_dir, f"top_{top_n}_analysis.xlsx")

    print(f"\n5. Writing report (top {top_n})...")
    if HAS_OPENPYXL:
        write_excel(results_by_vina, results_by_le, metadata, output_path, top_n)
    else:
        print("  openpyxl not available, writing TSV files instead")
        write_tsv_fallback(results_by_vina, results_by_le, metadata, output_dir, top_n)

    print("\n" + "=" * 60)
    print("ANALYSIS COMPLETE")
    print("=" * 60)


if __name__ == "__main__":
    main()
